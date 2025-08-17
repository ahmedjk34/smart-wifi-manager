import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
import logging
from pathlib import Path
from typing import List, Dict, Optional

# Set up logging (will be configured in __init__)
logger = None
logger = logging.getLogger(__name__)

class SmartV1BenchmarkAnalyzer:
    """Analyzer for SmartWifiManagerV1 benchmark data with improved logging and structure."""
    
    def __init__(self, csv_filename: str = 'smartv1-benchmark.csv'):
        self.csv_filename = csv_filename
        self.base_dir = Path(__file__).parent
        self.results_dir = self.base_dir / 'test_results' / 'smartv1_test'
        self.df: Optional[pd.DataFrame] = None
        self.plots_generated: List[str] = []
        
        # Create results directory first
        self._create_results_directory()
        
        # Set up logging after creating results directory
        self._setup_logging()
        
    def _setup_logging(self) -> None:
        """Set up logging configuration with proper file path."""
        global logger
        
        # Create log file in the results directory
        log_file = self.results_dir / 'smartv1_analysis.log'
        
        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, mode='w'),  # Overwrite log file each run
                logging.StreamHandler()
            ],
            force=True  # Override any existing configuration
        )
        
        logger = logging.getLogger(__name__)
        logger.info(f"Logging initialized. Log file: {log_file}")
        
    def _create_results_directory(self) -> None:
        """Create the results directory structure."""
        try:
            self.results_dir.mkdir(parents=True, exist_ok=True)
            print(f"Results directory created/verified: {self.results_dir}")
        except Exception as e:
            print(f"Failed to create results directory: {e}")
            raise
    
    def load_data(self) -> bool:
        """Load CSV data with error handling."""
        # Try multiple possible locations for the CSV file
        possible_paths = [
            self.base_dir / self.csv_filename,  # Same directory as script
            self.base_dir.parent / self.csv_filename,  # Parent directory
            self.base_dir / '..' / self.csv_filename,  # Parent directory (alternative)
        ]
        
        csv_path = None
        for path in possible_paths:
            if path.exists():
                csv_path = path
                break
        
        if csv_path is None:
            logger.error(f"CSV file '{self.csv_filename}' not found in any of these locations:")
            for path in possible_paths:
                logger.error(f"  - {path}")
            return False
        
        try:    
            self.df = pd.read_csv(csv_path)
            logger.info(f"Successfully loaded {len(self.df)} rows from {csv_path}")
            
            # Log basic data information
            logger.info(f"Columns found: {list(self.df.columns)}")
            logger.info(f"Data shape: {self.df.shape}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error loading CSV data: {e}")
            return False
    
    def preprocess_data(self) -> None:
        """Preprocess data for analysis."""
        if self.df is None:
            logger.error("No data loaded. Call load_data() first.")
            return
        
        # Create standardized column mappings
        column_mappings = {
            'Distance': 'distance',
            'Speed': 'speed', 
            'Interferers': 'interferers',
            'PacketSize': 'packet_size',
            'TrafficRate': 'traffic_rate',
            'Throughput(Mbps)': 'throughput',
            'PacketLoss(%)': 'packet_loss',
            'AvgDelay(ms)': 'avg_delay',
            'RxPackets': 'rx_packets',
            'TxPackets': 'tx_packets'
        }
        
        # Rename columns to standardized names
        self.df = self.df.rename(columns=column_mappings)
        logger.info(f"Renamed columns. New columns: {list(self.df.columns)}")
        
        # Convert traffic_rate to numeric if it exists
        if 'traffic_rate' in self.df.columns:
            try:
                self.df['traffic_rate_num'] = (
                    self.df['traffic_rate']
                    .astype(str)
                    .str.replace('Mbps', '', case=False)
                    .str.replace(' ', '')
                    .astype(float)
                )
                logger.info("Successfully converted traffic_rate to numeric")
            except Exception as e:
                logger.warning(f"Could not convert traffic_rate to numeric: {e}")
                self.df['traffic_rate_num'] = None
        else:
            self.df['traffic_rate_num'] = None
            logger.info("No traffic_rate column found")
    
    def display_summary(self) -> None:
        """Display data summary with logging."""
        if self.df is None:
            logger.error("No data loaded.")
            return
            
        logger.info("=== DATA SUMMARY ===")
        print("\nFirst few rows:")
        print(self.df.head())
        
        print(f"\nColumns after preprocessing ({len(self.df.columns)}):")
        print(self.df.columns.tolist())
        
        print("\nData types:")
        print(self.df.dtypes)
        
        print("\nSummary statistics:")
        print(self.df.describe(include='all'))
        
        # Check for missing values
        missing_values = self.df.isnull().sum()
        if missing_values.any():
            print("\nMissing values:")
            print(missing_values[missing_values > 0])
        else:
            logger.info("No missing values found")
    
    def _create_plot(self, plot_func, filename: str, title: str) -> bool:
        """Helper function to create and save plots with error handling."""
        try:
            logger.info(f"Creating plot: {filename}")
            
            # Set style for better looking plots
            plt.style.use('default')
            
            fig, ax = plt.subplots(figsize=(10, 6))
            plot_func(ax)
            
            plt.title(title, fontsize=14, fontweight='bold')
            plt.tight_layout()
            
            plot_path = self.results_dir / f"{filename}.png"
            fig.savefig(plot_path, bbox_inches='tight', dpi=300, facecolor='white')
            plt.close(fig)
            
            # Verify file was created
            if plot_path.exists():
                file_size = plot_path.stat().st_size
                logger.info(f"Plot saved successfully: {plot_path} (Size: {file_size} bytes)")
                self.plots_generated.append(f"{filename}.png")
                return True
            else:
                logger.error(f"Plot file was not created: {plot_path}")
                return False
            
        except Exception as e:
            logger.error(f"Error creating plot {filename}: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return False
    
    def create_throughput_analysis_plots(self) -> None:
        """Create all throughput analysis plots."""
        if self.df is None:
            logger.warning("Cannot create throughput plots - no data loaded")
            return
            
        if 'throughput' not in self.df.columns:
            logger.warning("Cannot create throughput plots - missing throughput column")
            return
        
        logger.info("Creating throughput analysis plots...")
        
        # Throughput vs Distance
        if 'distance' in self.df.columns:
            def plot_distance(ax):
                sns.boxplot(data=self.df, x='distance', y='throughput', ax=ax)
                ax.set_xlabel('Distance (m)')
                ax.set_ylabel('Throughput (Mbps)')
                ax.tick_params(axis='x', rotation=45)
            
            success = self._create_plot(plot_distance, 'throughput_vs_distance', 'Throughput vs Distance')
            logger.info(f"Distance plot result: {'Success' if success else 'Failed'}")
        else:
            logger.info("Skipping distance plot - no distance column")
        
        # Throughput vs Traffic Rate
        if 'traffic_rate' in self.df.columns:
            def plot_traffic(ax):
                sns.boxplot(data=self.df, x='traffic_rate', y='throughput', ax=ax)
                ax.set_xlabel('Traffic Rate')
                ax.set_ylabel('Throughput (Mbps)')
                ax.tick_params(axis='x', rotation=45)
            
            success = self._create_plot(plot_traffic, 'throughput_vs_traffic_rate', 'Throughput vs Traffic Rate')
            logger.info(f"Traffic rate plot result: {'Success' if success else 'Failed'}")
        else:
            logger.info("Skipping traffic rate plot - no traffic_rate column")
        
        # Throughput vs Interferers
        if 'interferers' in self.df.columns:
            def plot_interferers(ax):
                sns.boxplot(data=self.df, x='interferers', y='throughput', ax=ax)
                ax.set_xlabel('Number of Interferers')
                ax.set_ylabel('Throughput (Mbps)')
            
            success = self._create_plot(plot_interferers, 'throughput_vs_interferers', 'Throughput vs Interferers')
            logger.info(f"Interferers plot result: {'Success' if success else 'Failed'}")
        else:
            logger.info("Skipping interferers plot - no interferers column")
        
        # Throughput vs Speed
        if 'speed' in self.df.columns:
            def plot_speed(ax):
                sns.boxplot(data=self.df, x='speed', y='throughput', ax=ax)
                ax.set_xlabel('Speed (m/s)')
                ax.set_ylabel('Throughput (Mbps)')
                ax.tick_params(axis='x', rotation=45)
            
            success = self._create_plot(plot_speed, 'throughput_vs_speed', 'Throughput vs Speed')
            logger.info(f"Speed plot result: {'Success' if success else 'Failed'}")
        else:
            logger.info("Skipping speed plot - no speed column")
        
        # Throughput vs Packet Size
        if 'packet_size' in self.df.columns:
            def plot_packet_size(ax):
                sns.boxplot(data=self.df, x='packet_size', y='throughput', ax=ax)
                ax.set_xlabel('Packet Size (bytes)')
                ax.set_ylabel('Throughput (Mbps)')
                ax.tick_params(axis='x', rotation=45)
            
            success = self._create_plot(plot_packet_size, 'throughput_vs_packet_size', 'Throughput vs Packet Size')
            logger.info(f"Packet size plot result: {'Success' if success else 'Failed'}")
        else:
            logger.info("Skipping packet size plot - no packet_size column")
    
    def create_pairplot(self) -> None:
        """Create pairplot for all numeric variables."""
        if self.df is None:
            logger.warning("No data available for pairplot")
            return
        
        # Select numeric columns for pairplot
        numeric_cols = ['distance', 'speed', 'interferers', 'packet_size', 'traffic_rate_num', 'throughput']
        available_cols = [col for col in numeric_cols if col in self.df.columns and self.df[col].dtype in ['int64', 'float64']]
        
        logger.info(f"Available numeric columns: {available_cols}")
        
        if len(available_cols) < 2:
            logger.warning(f"Not enough numeric columns for pairplot. Available: {available_cols}")
            return
        
        try:
            logger.info(f"Creating pairplot with columns: {available_cols}")
            
            # Create pairplot
            g = sns.pairplot(self.df[available_cols], diag_kind='kde')
            g.fig.suptitle('Pairwise Relationships Between Variables', y=1.02)
            
            pairplot_path = self.results_dir / "pairplot.png"
            g.savefig(pairplot_path, bbox_inches='tight', dpi=300, facecolor='white')
            plt.close(g.fig)
            
            # Verify file was created
            if pairplot_path.exists():
                file_size = pairplot_path.stat().st_size
                logger.info(f"Pairplot saved successfully: {pairplot_path} (Size: {file_size} bytes)")
                self.plots_generated.append('pairplot.png')
            else:
                logger.error(f"Pairplot file was not created: {pairplot_path}")
            
        except Exception as e:
            logger.error(f"Error creating pairplot: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
        
        logger.info(f"Total plots generated so far: {len(self.plots_generated)}")
    
    def calculate_group_summaries(self) -> Dict[str, pd.DataFrame]:
        """Calculate group summaries for each parameter."""
        if self.df is None or 'throughput' not in self.df.columns:
            logger.warning("Cannot calculate group summaries - missing data or throughput column")
            return {}
        
        group_params = ['distance', 'speed', 'interferers', 'packet_size', 'traffic_rate']
        group_summaries = {}
        
        logger.info("Calculating group summaries...")
        
        for param in group_params:
            if param in self.df.columns:
                try:
                    summary = (
                        self.df.groupby(param)['throughput']
                        .agg(['mean', 'std', 'count'])
                        .round(3)
                        .reset_index()
                    )
                    summary.columns = [param, 'mean_throughput', 'std_throughput', 'count']
                    group_summaries[param] = summary
                    
                    logger.info(f"Summary calculated for {param}:")
                    print(f"\nMean throughput by {param}:")
                    print(summary)
                    
                except Exception as e:
                    logger.error(f"Error calculating summary for {param}: {e}")
        
        return group_summaries
    
    def export_to_excel(self, group_summaries: Dict[str, pd.DataFrame]) -> None:
        """Export results to Excel with formatting."""
        excel_path = self.results_dir / 'smartv1_benchmark_results.xlsx'
        
        try:
            logger.info(f"Exporting results to Excel: {excel_path}")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Raw data
                self.df.to_excel(writer, sheet_name='RawData', index=False)
                
                # Group summaries
                for param, summary_df in group_summaries.items():
                    sheet_name = f'Summary_{param}'
                    summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Analysis summary
                analysis_summary = pd.DataFrame({
                    'Metric': ['Total Records', 'Columns', 'Plots Generated'],
                    'Value': [len(self.df), len(self.df.columns), len(self.plots_generated)]
                })
                analysis_summary.to_excel(writer, sheet_name='Analysis_Info', index=False)
            
            # Apply formatting
            self._format_excel_file(excel_path)
            logger.info(f"Excel file exported successfully: {excel_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
    
    def _format_excel_file(self, excel_path: Path) -> None:
        """Apply formatting to Excel file."""
        try:
            wb = load_workbook(excel_path)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Bold and center headers
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                
                # Auto-fit columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            logger.info("Excel formatting applied successfully")
            
        except Exception as e:
            logger.error(f"Error formatting Excel file: {e}")
    
    def run_complete_analysis(self) -> None:
        """Run the complete analysis pipeline."""
        logger.info("=== Starting SmartWifiManagerV1 Benchmark Analysis ===")
        
        # Load and preprocess data
        if not self.load_data():
            logger.error("Analysis failed - could not load data")
            return
        
        self.preprocess_data()
        self.display_summary()
        
        # Generate plots
        self.create_throughput_analysis_plots()
        self.create_pairplot()
        
        # Calculate summaries and export
        group_summaries = self.calculate_group_summaries()
        self.export_to_excel(group_summaries)
        
        # Final summary
        logger.info("=== Analysis Complete ===")
        logger.info(f"Results saved in: {self.results_dir}")
        logger.info(f"Plots generated: {len(self.plots_generated)}")
        logger.info(f"Generated files: {self.plots_generated + ['smartv1_benchmark_results.xlsx']}")
        
        print(f"\n{'='*50}")
        print(f"ANALYSIS COMPLETE")
        print(f"{'='*50}")
        print(f"Results directory: {self.results_dir}")
        print(f"Excel file: smartv1_benchmark_results.xlsx")
        print(f"Plots generated: {len(self.plots_generated)}")
        print(f"Log file: {self.results_dir}/smartv1_analysis.log")


def main():
    """Main function to run the analysis."""
    try:
        # Initialize analyzer (this will set up logging)
        analyzer = SmartV1BenchmarkAnalyzer()
        
        # Run complete analysis
        analyzer.run_complete_analysis()
        
    except Exception as e:
        if logger:
            logger.error(f"Analysis failed with error: {e}")
        else:
            print(f"Analysis failed with error: {e}")
        raise


if __name__ == "__main__":
    main()